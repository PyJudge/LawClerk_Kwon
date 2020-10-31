#%%
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import chain

src = load_workbook("사실관계 정리표.xlsx")
form = load_workbook("temp_template.xlsx")
src_ws = src.active
form_ws = form.active

# %%
end_n_row = ""
for i, row in enumerate(src_ws.iter_cols(min_row = 2)): # omit first row 
    for j, cell in enumerate(row):
        position = [i+2, j+2]         # where to start
        print(*position, n_cell(position), form_ws[n_cell(position)].value)
        if n_cell(position)[0] == "A":
            input()
        form_ws[n_cell(position)] = cell.value
        end_n_row = n_cell(position)[1:]

form_ws.move_range(cell_range = "B2:B{}".format(end_n_row), cols = -1)

for i in range(2, int(end_n_row) + 1):
    form_ws[n_cell([2, i])] = '=IF(SUMPRODUCT(1÷COUNTIF($A$2:A{},A{}))=1,A{},"")'.format(i, i, i)

form.save("일시.xlsx")

# %%
# 11 > A1
# 56 > E6
def n_cell(pos: list()) -> str:
    return chr(pos[0] + 64) + str(pos[1])
# %%

# %%
# 새로운 인텔리 버젼 ! 
# 1단계 sentence를 발라낸다 
# 2단계 sentence 중, 증거가 있는 것을 찾는다 
# 2-1단계  
# 2-1단계 증거가 많은 것을 우선한다 
# 
# 
raw_data = [
    [1, 2, 3, 4, 5, 6,], [2, 3, 6, 7], [4, 5, 8], [1, 6, 7, 9], [3, 6], [2, 6, 9], 
    [5, 1, 2, 3, 9], [3, 6, 8]]

# %%
remaining = set([raw for raw in raw_data])
#%%
# %%
