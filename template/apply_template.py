#%%
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import chain
import os 


def n_cell(pos: list()) -> str:
    return chr(pos[0] + 64) + str(pos[1])

def apply_template(PDF_dir: str, file: str = "사실관계 정리표.xlsx") -> None:
    src = load_workbook(os.path.join(PDF_dir, file))
    form = load_workbook(os.path.join(os.getcwd(), "template", "temp_template.xlsx"))
    src_ws = src.active
    form_ws = form.active

    # %%
    end_n_row = ""
    for i, row in enumerate(src_ws.iter_cols(min_row = 2)): # omit first row 
        for j, cell in enumerate(row):
            position = [i+2, j+2]         # where to start
            print(*position, n_cell(position), form_ws[n_cell(position)].value)
            if n_cell(position)[0] == "A":
                input()
            form_ws[n_cell(position)] = cell.value
            end_n_row = n_cell(position)[1:]

    form_ws.move_range(cell_range = "B2:B{}".format(end_n_row), cols = -1)

    for i in range(2, int(end_n_row) + 1):
        form_ws[n_cell([2, i])] = '=IF(SUMPRODUCT(1÷COUNTIF($A$2:A{},A{}))=1,A{},"")'.format(i, i, i)

    form.save(os.path.join(PDF_dir, "사실관계 정리표-서식 적용.xlsx"))

# %%
# 11 > A1
# 56 > E6
# %%
